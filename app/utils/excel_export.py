import io
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.models.financial_record import RecordType

def generate_excel_bytes(records_iterator) -> bytes:
    """Generates an aesthetically formatted Excel file with a summary at the bottom."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Financial Records"

    # Define Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    
    headers = ["ID", "User ID", "Amount", "Type", "Category", "Date", "Notes", "Created At"]
    
    # Write Headers
    ws.append(headers)
    for col_num, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # Set column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 30
    ws.column_dimensions['H'].width = 25

    total_income = Decimal("0.0")
    total_expense = Decimal("0.0")

    row_num = 2
    for record in records_iterator:
        amount = record.amount or Decimal("0.0")
        rtype = record.type.value if record.type else ""
        
        if record.type == RecordType.INCOME:
            total_income += amount
        elif record.type == RecordType.EXPENSE:
            total_expense += amount

        ws.append([
            record.id,
            record.user_id,
            float(amount),
            rtype,
            record.category,
            str(record.date),
            record.notes or "",
            str(record.created_at)[:19]  # truncate milliseconds for clean look
        ])
        
        # Color coding amount/type
        if rtype == "INCOME":
            ws.cell(row=row_num, column=4).font = Font(color="008000", bold=True)
            ws.cell(row=row_num, column=3).font = Font(color="008000")
        elif rtype == "EXPENSE":
            ws.cell(row=row_num, column=4).font = Font(color="FF0000", bold=True)
            ws.cell(row=row_num, column=3).font = Font(color="FF0000")
            
        row_num += 1

    # Add Totals at the bottom
    row_num += 2
    # Add a summary section
    ws.cell(row=row_num, column=2, value="SUMMARY REPORT").font = Font(bold=True)
    
    row_num += 1
    ws.cell(row=row_num, column=2, value="Total Income")
    ws.cell(row=row_num, column=3, value=float(total_income)).font = Font(color="008000", bold=True)
    
    row_num += 1
    ws.cell(row=row_num, column=2, value="Total Expense")
    ws.cell(row=row_num, column=3, value=float(total_expense)).font = Font(color="FF0000", bold=True)
    
    row_num += 1
    net = total_income - total_expense
    ws.cell(row=row_num, column=2, value="Net Balance").font = Font(bold=True)
    
    net_color = "008000" if net >= 0 else "FF0000"
    ws.cell(row=row_num, column=3, value=float(net)).font = Font(color=net_color, bold=True)
    
    # Save to buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()
